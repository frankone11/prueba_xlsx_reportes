from openpyxl import load_workbook

def prueba_reportes():
	workbook = load_workbook(filename='reportes.xlsx')
	
	sheet = workbook.active
	i = 0
	while True:
		if i == 10:
			break
		sheet.cell(row=i+5, column=1).value = str(i+1)
		sheet.cell(row=i+5, column=2).value = "RFC" + str(i+1)
		sheet.cell(row=i+5, column=3).value = "Descripción de " + str(i+1)
		print("Cell " + str(i+1))
		i+=1
	workbook.save('reportes_1.xlsx')
	return 0

def prueba_reporte():
	workbook = load_workbook(filename='reporte.xlsx')
	
	sheet = workbook.active
	i = 0
	while True:
		if i == 10:
			break
		sheet.cell(row=i+5, column=1).value = str(i+1)
		sheet.cell(row=i+5, column=2).value = "RFC" + str(i+1)
		sheet.cell(row=i+5, column=4).value = "Descripción de " + str(i+1)
		print("Cell " + str(i+1))
		i+=1
	workbook.save('reporte_1.xlsx')
	return 0